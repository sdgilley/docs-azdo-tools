# Updates Excel spreadsheet 
# Queries from a Work Item column, updates the State and AssignedTo columns in the file
# Preserves formulas in the file
# REQUIRES the excel file to have General permissions, otherwise it will give an error

# !IMPORTANT - sign in with az login --use-device-code before running this script
import pandas as pd
from openpyxl import load_workbook
import os
import sys

# Add the parent directory to the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import helpers.azdo as a


def update_spreadsheet(file_path,tab):
    df = pd.read_excel(file_path, sheet_name=tab)
    # sort the DataFrame by the 'Work Item' column
    df = df.sort_values(by='Work Item')  
    print(f"Size of DataFrame: {df.shape}")

    # Ensure Work Item column is of string type before using .str.replace
    df['Work Item'] = df['Work Item'].astype(str)
    # get the real work items, not the blank or nan ones
    query_df = df[df['Work Item'].notna() & (df['Work Item'] != 'nan')].copy()

    # Ensure Work Item column is of string type for the query and remove floating-point artifacts
    query_df['Work Item'] = query_df['Work Item'].astype(str).str.split('.').str[0]


    # Define the Azure DevOps query
    project_name = "Content"
    columns = ['System.Id', 'System.State', 'System.AssignedTo']
    query = f"""
        SELECT {','.join(columns)}
        FROM workitems 
        WHERE [System.TeamProject] = 'Content' AND [System.Id] IN ({','.join(query_df['Work Item'])})
        """
    # Query work items
    work_items = a.query_work_items(query, columns)
    print(work_items.columns)
    print(work_items)

    # Convert the Id column in work_items to string to match the Work Item column in query_df
    work_items['Id'] = work_items['Id'].astype(str)
    df['Work Item'] = df['Work Item'].astype(str).str.split('.').str[0]
    # Merge the updated work items into the original DataFrame based on Work Item and Id
    df = df.merge(
        work_items.rename(columns={'Id': 'Work Item'}),
        on='Work Item',
        how='left',
        suffixes=('', '_updated')  # Add suffix to avoid overwriting columns immediately
    )

    # Replace the existing State and AssignedTo columns with the updated values
    df['State'] = df['State_updated']  # Get latest state
    df['AssignedTo'] = df['AssignedTo_updated']  # get latest AssignedTo
    # if work item is nan, set to blank
    df['Work Item'] = df['Work Item'].fillna('')

    # Drop the temporary columns created by the merge
    df = df.drop(columns=['State_updated', 'AssignedTo_updated'])
    print(f"Updated DataFrame size: {df.shape}")

    # Save the updated DataFrame back to the Excel file
    # Load the original workbook with openpyxl
    wb = load_workbook(file_path)
    ws = wb[tab]  # Specify the tab to update

    # Remove any filters applied to the worksheet
    ws.auto_filter.ref = None  # Clear any existing filters

    # Preserve all formulas in the worksheet
    formulas = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Start from row 2 to skip the header
        for cell in row:
            if cell.data_type == "f":  # Check if the cell contains a formula
                formulas[(cell.row, cell.column)] = cell.value  # Store the formula with its row and column

    # Write the updated data back to the workbook
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx + 2, column=col_idx, value=value)  # +2 to account for header row

    # Restore the preserved formulas
    for (row_idx, col_idx), formula in formulas.items():
        ws.cell(row=row_idx, column=col_idx, value=formula)

    # Save the workbook
    wb.save(file_path)
    print(f"Updated spreadsheet saved to {file_path}")

if __name__ == "__main__":
    
    # Local file path to the Excel file
    file_path = r"C:\Users\sgilley\OneDrive - Microsoft\AI Foundry\doc-updates-build2025.xlsx"
    tab = "ai-services"
    update_spreadsheet(file_path, tab)
# storing here in case it gets mangled:
# A=HYPERLINK(CONCAT("https://review.learn.microsoft.com/azure/",SUBSTITUTE([@[File Path]],".md",""),"?branch=release-build-ai-foundry"),"🔗")
# J=IF(NOT([@[Work Item]]="nan"),HYPERLINK(CONCAT("https://dev.azure.com/msft-skilling/Content/_queries/edit/",[@[Work Item]]),"🔗"),"")