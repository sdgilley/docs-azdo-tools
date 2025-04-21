import pandas as pd
from openpyxl import load_workbook
import helpers.azdo as a

# Local file path to the Excel file
file_path = r"C:\Users\sgilley\OneDrive - Microsoft\AI Foundry\doc-updates-build2025-copy2.xlsx"
tab="WorkItems"  # Specify the tab name here
df = pd.read_excel(file_path, sheet_name=tab)
# sort the DataFrame by the 'Work Item' column
df = df.sort_values(by='Work Item')  
print(f"Size of DataFrame: {df.shape}")

# Ensure Work Item column is of string type before using .str.replace
df['Work Item'] = df['Work Item'].astype(str)
query_df = df[df['Work Item'].notna() & (df['Work Item'] != 'nan')].copy()

# Ensure Work Item column is of string type for the query and remove floating-point artifacts
query_df['Work Item'] = query_df['Work Item'].astype(str).str.split('.').str[0]


# Define the Azure DevOps query
project_name = "Content"
columns = ['System.Id', 'System.State', 'System.AssignedTo']
query = f"""
    SELECT {','.join(columns)}
    FROM workitems 
    WHERE [System.TeamProject] = 'Content' AND [System.Id] IN ({','.join(query_df['Work Item'])})
    """
# Query work items
work_items = a.query_work_items(query, columns)
print(work_items.columns)

# Convert the Id column in work_items to string to match the Work Item column in query_df
work_items['Id'] = work_items['Id'].astype(str)
df['Work Item'] = df['Work Item'].astype(str).str.split('.').str[0]
# Merge the updated work items into the original DataFrame based on Work Item and Id
df = df.merge(
    work_items.rename(columns={'Id': 'Work Item'}),
    on='Work Item',
    how='left',
    suffixes=('', '_updated')  # Add suffix to avoid overwriting columns immediately
)

# Replace the existing State and AssignedTo columns with the updated values
df['State'] = df['State_updated']  # Get latest state
df['AssignedTo'] = df['AssignedTo_updated']  # get latest AssignedTo
# if work item is nan, set to blank
df['Work Item'] = df['Work Item'].fillna('')

# Drop the temporary columns created by the merge
df = df.drop(columns=['State_updated', 'AssignedTo_updated'])
print(f"Updated DataFrame size: {df.shape}")

# Load the original workbook with openpyxl
wb = load_workbook(file_path)
ws = wb[tab]  # Specify the tab to update

# Remove any filters applied to the worksheet
ws.auto_filter.ref = None  # Clear any existing filters

# Write the updated data back to the workbook
for row_idx, row in df.iterrows():
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx + 2, column=col_idx, value=value)  # +2 to account for header row

# Preserve workbook properties
wb.properties = wb.properties  # Retain original workbook properties
wb.security = wb.security      # Retain original workbook security settings

# Save the workbook
wb.save(file_path)
print(f"Updated spreadsheet saved to {file_path}")

# storing here for now...
# =HYPERLINK(CONCAT("https://review.learn.microsoft.com/azure/",[@[File Path]],"?branch=release-build-ai-foundry"))
# =HYPERLINK(CONCAT("https://review.learn.microsoft.com/azure/",SUBSTITUTE([@[File Path]],"md",""),"?branch=release-build-ai-foundry"))